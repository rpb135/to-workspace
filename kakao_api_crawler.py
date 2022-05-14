import requests, json
import openpyxl

api_headers = {
    'authority': 'search.map.kakao.com',
    'scheme': 'https',
    'accept': '*/*',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
    'cookie': 'webid=799131c21c01486d9898e5ae2c1ea860; webid_ts=1576916209172; TIARA=hfIvEE.j8log9bSYDUKC3gHtx5HbPPfRxEV_SIgVzTb9hNqp1BZGnDBs.3.-heKJAHot8EmDyxrfsZH2Fa8AZ7WjL11rfV6k; _kadu=HUIeEfDkMrFMVz0e_1624844124186; _kptid=f0c1a3ce716646539ca13b903f23a470; _kp_collector=KP.608151197.1630323919768; _ga=GA1.2.1653037846.1606881786; _ga_80D69HE0QD=GS1.1.1630692280.10.1.1630693065.0; _gcl_au=1.1.45959106.1651232501; _fbp=fb.1.1651232501504.90169813; __T_=1; _kawlt=7F7T1xnQqnlZ683fsl1NAKZJh_PdTCsEgjdgJ4ZdbtKTyoY2t2KxHGr5zCn8ymDkYzHHQriymX3jmzPKeCW_O8RH7JA0XaI0DLL_6cCs1fWf0Q-tvS_sk7rx62Pnp7Gs; _kawltea=1651782964; _karmt=dlO2FCbqVGLnantFiBPIuDPaAWfSNwFgTVnK5Ip6galxxvs_I0SHYMW1KySUudWz; _karmtea=1651793764; _kahai=34e4407126892284a87f6744530b0b9463475cd03a8697664ed62900e5947281; _T_ANO=CPgG4DoD5kyf8zLReWXwzaWQdxTPycqoLee544f9P5gNDGIpiv4JdIOpCBD9SxkqkNrvkSz60QiVZobRqaewC7mOFgC4O/cVDFHuwBIm/+3X64WT6scxs8RPhr9y2b5pXutW8s5yVwnzmo9m8Nc6Q8zu2bT6Yr7FAF+XTJmxvGNfBQ+YtZA8NvKgBBYKRtQZJ5tGjz4p5Tv1dcuVvCUkYHwXnDKZMAhnXrIxnHmegQ/G6wgP6fIc4KXiNSVuFaLl+C+hBR6SQLVBt8/3jhdTGDUT8Lh9/xbUhgHg2QX3BotL14BJp538BvEsi2ilqSp1ymVolLLj5qxZJxB56kMorA==',
    'referer': 'https://map.kakao.com/',
    'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="101", "Google Chrome";v="101"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'script',
    'sec-fetch-mode': 'no-cors',
    'sec-fetch-site': 'same-site',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36'
}

new_headers = {
    'Content-Type': 'application/json; charset=utf-8',
    'accept': '*/*',
    'cookie': 'webid=799131c21c01486d9898e5ae2c1ea860; webid_ts=1576916209172; TIARA=hfIvEE.j8log9bSYDUKC3gHtx5HbPPfRxEV_SIgVzTb9hNqp1BZGnDBs.3.-heKJAHot8EmDyxrfsZH2Fa8AZ7WjL11rfV6k; _kadu=HUIeEfDkMrFMVz0e_1624844124186; _kptid=f0c1a3ce716646539ca13b903f23a470; _kp_collector=KP.608151197.1630323919768; _ga=GA1.2.1653037846.1606881786; _ga_80D69HE0QD=GS1.1.1630692280.10.1.1630693065.0; _gcl_au=1.1.45959106.1651232501; _fbp=fb.1.1651232501504.90169813; __T_=1; _kawlt=7F7T1xnQqnlZ683fsl1NAKZJh_PdTCsEgjdgJ4ZdbtKTyoY2t2KxHGr5zCn8ymDkYzHHQriymX3jmzPKeCW_O8RH7JA0XaI0DLL_6cCs1fWf0Q-tvS_sk7rx62Pnp7Gs; _kawltea=1651782964; _karmt=dlO2FCbqVGLnantFiBPIuDPaAWfSNwFgTVnK5Ip6galxxvs_I0SHYMW1KySUudWz; _karmtea=1651793764; _kahai=34e4407126892284a87f6744530b0b9463475cd03a8697664ed62900e5947281; _T_ANO=CPgG4DoD5kyf8zLReWXwzaWQdxTPycqoLee544f9P5gNDGIpiv4JdIOpCBD9SxkqkNrvkSz60QiVZobRqaewC7mOFgC4O/cVDFHuwBIm/+3X64WT6scxs8RPhr9y2b5pXutW8s5yVwnzmo9m8Nc6Q8zu2bT6Yr7FAF+XTJmxvGNfBQ+YtZA8NvKgBBYKRtQZJ5tGjz4p5Tv1dcuVvCUkYHwXnDKZMAhnXrIxnHmegQ/G6wgP6fIc4KXiNSVuFaLl+C+hBR6SQLVBt8/3jhdTGDUT8Lh9/xbUhgHg2QX3BotL14BJp538BvEsi2ilqSp1ymVolLLj5qxZJxB56kMorA==',
    'referer': 'https://map.kakao.com/',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36'
}


def search_list(search_word, page_nm):
    return 'https://search.map.kakao.com/mapsearch/map.daum?callback=jQuery18105813139345637819_1651706709560&q='+ search_word +'&msFlag=S&page='+ str(page_nm) +'&sort=0'


def response_obj(url, new_headers, df_dict):

    with requests.get(url, headers=new_headers) as res:

        a = res.text.index('(')
        #b = res.text.index(')')
        search_word = res.text[a+1:-2]
        json_obj = json.loads(search_word)
        obj_len = len(json_obj['place'])

        if obj_len == 0:
            #검색결과가 없으므로 종료
            #print(obj_len)
            df_dict['__End Point__'] = 0
            return df_dict
        else:
            for i in range(obj_len):
                print(str(i) + ' 번째')
                print(json_obj['place'][i]['address'])
                df_dict[name(json_obj['place'][i]['name'])] =  json_obj['place'][i]['address']

        res.close
        res.close()
        return df_dict

class name(object):
    def __init__(self, name):
        self.name = name

    def __str__(self):
        return self.name

    def __repr__(self):
        return "'"+self.name+"'"


def main():

    search_keyword = input('검색어 입력 : ')
    df_dict = {}
    wb = openpyxl.Workbook()
    sht = wb.create_sheet('1', 0)
    sht.cell(column=1, row=1, value='이름')
    sht.cell(column=2, row=1, value='주소')

    for page_i in range(1, 500):
        url = search_list(search_keyword, page_i)
        print(page_i)
        df_dict = response_obj(url, new_headers, df_dict)
        if '__End Point__' in df_dict:
            print('---PASS---')
            break

    if '__End Point__' in df_dict:
        del(df_dict['__End Point__'])
        #print('---종료---')

    for dict_i, (key, elem) in enumerate(df_dict.items()):
        print(dict_i, key, elem)
        sht.cell(column=1, row=dict_i+2, value=str(key))
        sht.cell(column=2, row=dict_i+2, value=elem)

    wb.save(search_keyword + '_리스트.xlsx')


if __name__ == "__main__":
    main()
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Border, Side, Alignment

SHEET_MAX_NUM = 252 #2
SHEET_MAX_NUM2 = 435 #254

driver = webdriver.Chrome('./company_work/dbo_work/chromedriver')
#driver = webdriver.Chrome('./chromedriver')
driver.get('file:///C:/workspace/company_work/dbo_work/main.html')
driver.implicitly_wait(3)

#frame TREE 이동
element_TREE = driver.find_element_by_name('TREE')
driver.switch_to.frame(element_TREE)

#버튼 선택 > 목록 열기
btn1 = driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
btn2 = driver.find_element(By.XPATH, '//*[@id="btn-expand-nodes"]').click()
wb = openpyxl.load_workbook(filename='./company_work/dbo_work/table_define_1.xlsx')
#wb = openpyxl.load_workbook(filename='./company_work/dbo_work/table_define_2.xlsx')

for i in range(SHEET_MAX_NUM):
    print(i)
    
    #wb2 = openpyxl.load_workbook(filename='table_define_2.xlsx')
    #1번 시트 복사
    sht = wb['0']
    wb.copy_worksheet(sht)

    #driver.get('file:///C:/Users/User/OneDrive%20-%20%EC%8A%A4%ED%8C%8C%EC%9D%B4%EB%8D%94%ED%81%AC%EB%9E%98%ED%94%84%ED%8A%B8/%EB%B0%94%ED%83%95%20%ED%99%94%EB%A9%B4/%EC%97%85%EB%AC%B4/etc/dbo%EC%9E%91%EC%97%85/main.html')

    #table 목록 선택 (순차진행)
    list = driver.find_element(By.XPATH, '//*[@id="tree"]/ul/li['+ str(i+7) +']/a').click()
    #list = driver.find_element(By.XPATH, '//*[@id="tree"]/ul/li[+ str(i+252) +]/a').click()

    #초기 frame으로 이동
    driver.switch_to.default_content()

    #frame DATA 이동
    element_DATA = driver.find_element_by_name('DATA')
    driver.switch_to.frame(element_DATA)

    table_name = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[2]/div/div/h1').text
    print(table_name)

    #복사 된 0 Copy 시트명 변경
    ch_sht = wb['0 Copy']
    ch_sht.title = str(i+1)

    '''
    # Mapping Table init
    mapping_sht = wb1['Mapping Table']
    mapping_sht.cell(column=1, row=i+2, value=i+1)          #column : 1 고정, row : 변동 값, value : 시트명(반복횟수)
    mapping_sht.cell(column=2, row=i+2, value=table_name)   #column : 2 고정, row : 변동 값, value : TableName
    '''

    ws = wb[str(i+1)]
    ws.cell(column=3, row=4, value=table_name)
    ws.cell(column=3, row=4, value=table_name).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=3, row=4, value=table_name).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

    table = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table')
    tbody = table.find_element_by_tag_name('tbody')
    row = tbody.find_elements_by_tag_name('tr')

    for idx, tr in enumerate(row, start=0):

        if idx == 0:
            continue
        columns_name = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[2]').text
        columns_data_type = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) +  ']/td[3]').text
        columns_data_length = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) +  ']/td[4]').text
        columns_Not_Null_bl = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) +  ']/td[7]').text

        ws.cell(column=1, row=idx+6, value=idx)
        ws.cell(column=1, row=idx+6, value=idx).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=1, row=idx+6, value=idx).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=2, row=idx+6, value=columns_name)
        ws.cell(column=2, row=idx+6, value=columns_name).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=2, row=idx+6, value=columns_name).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        data_type = ''

        if columns_data_length == '':
            columns_data_length = 'MAX'
            data_type = columns_data_type + '(' + columns_data_length + ')'
            
        ws.cell(column=3, row=idx+6, value=data_type)
        ws.cell(column=3, row=idx+6, value=data_type).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=3, row=idx+6, value=data_type).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        if columns_Not_Null_bl == 'True':
            not_null = 'NOT NULL'
            ws.cell(column=4, row=idx+6, value=not_null)
        ws.cell(column=4, row=idx+6, value=not_null).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=4, row=idx+6, value=not_null).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        
        Idx_Key = ''
        Idx_text = ''
        try:
            key = ['Primary Key', 'Indexes', 'Cluster Key']

            key_icon = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[1]/div[1]').find_element_by_tag_name('svg').get_attribute('data-original-title')
            for i_key in key:
                if i_key in key_icon:
                    Idx_Key = i_key
                    Idx_text1 = key_icon.split(i_key+' ')
                    Idx_text = Idx_text1[1]

            key_icon = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[1]/div[2]').find_element_by_tag_name('svg').get_attribute('data-original-title')
            for i_key in key:
                if i_key in key_icon:
                    Idx_Key += ',' + i_key
                    Idx_text2 = key_icon.split(i_key+' ')
                    if Idx_text == Idx_text2[1]:
                        Idx_text = Idx_text
                    else:
                        Idx_text += ',' + Idx_text2[1]


            key_icon = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[1]/div[3]').find_element_by_tag_name('svg').get_attribute('data-original-title')
            for i_key in key:
                if i_key in key_icon:
                    Idx_Key += ',' + i_key
                    Idx_text3 = key_icon.split(i_key+' ')
                    if Idx_text == Idx_text3[1]:
                        Idx_text = Idx_text
                    else:
                        Idx_text += ',' + Idx_text3[1]


            columns_idx_name = driver.find_element(By.XPATH, '//*[@id="Indexes"]/div[2]/div/table/tbody/tr[2]/td[2]').text
        except:
            print('no have Key, PASS!')
        
        ws.cell(column=5, row=idx+6, value=Idx_Key)
        ws.cell(column=5, row=idx+6, value=Idx_Key).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=5, row=idx+6, value=Idx_Key).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=6, row=idx+6, value=Idx_text).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(column=6, row=idx+6, value=Idx_text).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=7, row=idx+6, value=Idx_text).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=8, row=idx+6, value=Idx_text).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.merge_cells(start_row=idx+6, start_column=6, end_row=idx+6, end_column=8)
        ws.cell(column=6, row=idx+6, value=Idx_text)

    #초기 frame으로 이동
    driver.switch_to.default_content()

    #frame TREE 이동
    element_TREE = driver.find_element_by_name('TREE')
    driver.switch_to.frame(element_TREE)


wb.save(filename='./company_work/dbo_work/table_define_1.xlsx')
driver.close()
#######################################################################################################################################

driver = webdriver.Chrome('./company_work/dbo_work/chromedriver')
#driver = webdriver.Chrome('./chromedriver')
driver.get('file:///C:/workspace/company_work/dbo_work/main.html')
driver.implicitly_wait(3)

#frame TREE 이동
element_TREE = driver.find_element_by_name('TREE')
driver.switch_to.frame(element_TREE)

#버튼 선택 > 목록 열기
btn1 = driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
btn2 = driver.find_element(By.XPATH, '//*[@id="btn-expand-nodes"]').click()
wb = openpyxl.load_workbook(filename='./company_work/dbo_work/table_define_2.xlsx')

for i in range(253, SHEET_MAX_NUM2):
    print(i)
    
    #wb2 = openpyxl.load_workbook(filename='table_define_2.xlsx')
    #1번 시트 복사
    sht = wb['0']
    wb.copy_worksheet(sht)

    #driver.get('file:///C:/Users/User/OneDrive%20-%20%EC%8A%A4%ED%8C%8C%EC%9D%B4%EB%8D%94%ED%81%AC%EB%9E%98%ED%94%84%ED%8A%B8/%EB%B0%94%ED%83%95%20%ED%99%94%EB%A9%B4/%EC%97%85%EB%AC%B4/etc/dbo%EC%9E%91%EC%97%85/main.html')

    #table 목록 선택 (순차진행)
    list = driver.find_element(By.XPATH, '//*[@id="tree"]/ul/li['+ str(i) +']/a').click()

    #초기 frame으로 이동
    driver.switch_to.default_content()

    #frame DATA 이동
    element_DATA = driver.find_element_by_name('DATA')
    driver.switch_to.frame(element_DATA)

    table_name = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[2]/div/div/h1').text
    print(table_name)

    #복사 된 1 Copy 시트명 변경
    ch_sht = wb['0 Copy']
    ch_sht.title = str(i)

    ws = wb[str(i)]
    ws.cell(column=3, row=4, value=table_name)
    ws.cell(column=3, row=4, value=table_name).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=3, row=4, value=table_name).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

    table = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table')
    tbody = table.find_element_by_tag_name('tbody')
    row = tbody.find_elements_by_tag_name('tr')

    for idx, tr in enumerate(row, start=0):

        if idx == 0:
            continue
        columns_name = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[2]').text
        columns_data_type = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) +  ']/td[3]').text
        columns_data_length = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) +  ']/td[4]').text
        columns_Not_Null_bl = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) +  ']/td[7]').text

        ws.cell(column=1, row=idx+6, value=idx)
        ws.cell(column=1, row=idx+6, value=idx).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=1, row=idx+6, value=idx).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=2, row=idx+6, value=columns_name)
        ws.cell(column=2, row=idx+6, value=columns_name).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=2, row=idx+6, value=columns_name).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        data_type = ''

        if columns_data_length == '':
            columns_data_length = 'MAX'
            data_type = columns_data_type + '(' + columns_data_length + ')'
            
        ws.cell(column=3, row=idx+6, value=data_type)
        ws.cell(column=3, row=idx+6, value=data_type).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=3, row=idx+6, value=data_type).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        if columns_Not_Null_bl == 'True':
            not_null = 'NOT NULL'
            ws.cell(column=4, row=idx+6, value=not_null)
        ws.cell(column=4, row=idx+6, value=not_null).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=4, row=idx+6, value=not_null).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        
        Idx_Key = ''
        Idx_text = ''
        try:
            key = ['Primary Key', 'Indexes', 'Cluster Key']

            key_icon = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[1]/div[1]').find_element_by_tag_name('svg').get_attribute('data-original-title')
            for i_key in key:
                if i_key in key_icon:
                    Idx_Key = i_key
                    Idx_text1 = key_icon.split(i_key+' ')
                    Idx_text = Idx_text1[1]

            key_icon = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[1]/div[2]').find_element_by_tag_name('svg').get_attribute('data-original-title')
            for i_key in key:
                if i_key in key_icon:
                    Idx_Key += ',' + i_key
                    Idx_text2 = key_icon.split(i_key+' ')
                    if Idx_text == Idx_text2[1]:
                        Idx_text = Idx_text
                    else:
                        Idx_text += ',' + Idx_text2[1]


            key_icon = driver.find_element(By.XPATH, '//*[@id="Columns"]/div[2]/div/table/tbody/tr[' + str(idx+1) + ']/td[1]/div[3]').find_element_by_tag_name('svg').get_attribute('data-original-title')
            for i_key in key:
                if i_key in key_icon:
                    Idx_Key += ',' + i_key
                    Idx_text3 = key_icon.split(i_key+' ')
                    if Idx_text == Idx_text3[1]:
                        Idx_text = Idx_text
                    else:
                        Idx_text += ',' +Idx_text3[1]


            columns_idx_name = driver.find_element(By.XPATH, '//*[@id="Indexes"]/div[2]/div/table/tbody/tr[2]/td[2]').text
        except:
            print('no have Key, PASS!')
        
        ws.cell(column=5, row=idx+6, value=Idx_Key)
        ws.cell(column=5, row=idx+6, value=Idx_Key).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(column=5, row=idx+6, value=Idx_Key).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=6, row=idx+6, value=Idx_text).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(column=6, row=idx+6, value=Idx_text).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=7, row=idx+6, value=Idx_text).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(column=8, row=idx+6, value=Idx_text).border = Border(left=Side(border_style='thin') ,right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.merge_cells(start_row=idx+6, start_column=6, end_row=idx+6, end_column=8)
        ws.cell(column=6, row=idx+6, value=Idx_text)

    #초기 frame으로 이동
    driver.switch_to.default_content()

    #frame TREE 이동
    element_TREE = driver.find_element_by_name('TREE')
    driver.switch_to.frame(element_TREE)


wb.save(filename='./company_work/dbo_work/table_define_2.xlsx')
driver.close()